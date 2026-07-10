"""
Home Loan Manager — Flask backend
SalaryBit (salarybit.in)

Mirrors the Tax Notice Shield / Insurance Mitra Railway deployment pattern:
- Deterministic rules engine (EMI math) in Python, no AI needed for this tool
- Flask + gunicorn, single service on Railway
- Razorpay order creation done server-side only (key_secret never touches the frontend)
- Leads appended as rows to an Excel file (leads.xlsx) — see /admin/leads to download

WHY THIS FILE ALSO HAS AN "EDUCATION" LAYER
---------------------------------------------
Every number the calculator produces (EMI, crossover point, total interest) is
neutral — it doesn't tell a first-time borrower what to DO with that number,
or what a bank will never bring up on its own. So this file also ships a
plain-language, borrower-first content layer covering the four moments where
common borrowers get hurt:
  1. Before you sign        -> hidden costs, bundled insurance, what to ask
  2. While you're repaying  -> why your EMI barely touches principal early on
  3. If you miss a payment  -> the real timeline (grace period -> NPA -> SARFAESI)
  4. If you're a guarantor  -> you are NOT a formality, you are a co-borrower

The tone throughout is deliberately "explain it like the reader has never
taken a loan before" — short sentences, rupee examples, no bank jargon left
unexplained. This is intentional: the audience is the common man, not a
finance professional.

Env vars required (set these in Railway, and locally in .env):
  RAZORPAY_KEY_ID
  RAZORPAY_KEY_SECRET
  LEADS_FILE_PATH            -> optional, defaults to ./leads.xlsx
                                 point this at a Railway Volume mount (e.g. /data/leads.xlsx)
                                 so leads survive redeploys — see README.
  RATES_FILE_PATH            -> optional, defaults to ./rates.json
                                 point this at the same Volume (e.g. /data/rates.json) so
                                 lender rate edits survive redeploys too — see README.
  ADMIN_TOKEN                -> secret used to protect the /admin/leads and /admin/rates routes
  REPORT_PRICE_PAISE         -> optional, defaults to 14900 (₹149)
  PORT                       -> set by Railway automatically
"""

import os
import json
import logging
import threading
import datetime

from flask import Flask, render_template, request, jsonify, send_file, abort, Response
from flask_cors import CORS
import razorpay
from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv

load_dotenv()

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("home-loan-manager")

app = Flask(__name__)
CORS(app)

# ── CONFIG ──────────────────────────────────────────────────────────────
RAZORPAY_KEY_ID = os.environ.get("RAZORPAY_KEY_ID", "")
RAZORPAY_KEY_SECRET = os.environ.get("RAZORPAY_KEY_SECRET", "")
REPORT_PRICE_PAISE = int(os.environ.get("REPORT_PRICE_PAISE", "14900"))  # ₹149

LEADS_FILE_PATH = os.environ.get("LEADS_FILE_PATH", os.path.join(os.path.dirname(__file__), "leads.xlsx"))
RATES_FILE_PATH = os.environ.get("RATES_FILE_PATH", os.path.join(os.path.dirname(__file__), "rates.json"))
ADMIN_TOKEN = os.environ.get("ADMIN_TOKEN", "")

LEADS_HEADERS = [
    "timestamp", "name", "phone", "email", "city",
    "loan_amount", "current_rate", "tenure_years",
    "best_lender", "total_savings",
    "payment_id", "order_id", "amount_paid_inr",
    "product", "status",
]

_leads_lock = threading.Lock()

razorpay_client = None
if RAZORPAY_KEY_ID and RAZORPAY_KEY_SECRET:
    razorpay_client = razorpay.Client(auth=(RAZORPAY_KEY_ID, RAZORPAY_KEY_SECRET))
else:
    logger.warning("Razorpay keys not set — /api/create-order and /api/verify-payment will fail until configured.")

# ── LENDER RATE DATA ─────────────────────────────────────────────────────
DEFAULT_LENDER_RATES = [
    {"name": "SBI (Public Sector)", "rate": 7.35},
    {"name": "PNB / Bank of Baroda", "rate": 7.55},
    {"name": "HDFC Bank", "rate": 7.90},
    {"name": "ICICI Bank", "rate": 8.10},
    {"name": "Kotak Mahindra Bank", "rate": 8.30},
    {"name": "Axis Bank", "rate": 8.60},
]

_rates_lock = threading.Lock()


def load_lender_rates() -> list:
    if os.path.exists(RATES_FILE_PATH):
        try:
            with open(RATES_FILE_PATH, "r") as f:
                data = json.load(f)
            if isinstance(data, list) and data:
                return data
        except Exception:
            logger.exception("Failed to read %s — falling back to defaults", RATES_FILE_PATH)
    return DEFAULT_LENDER_RATES


def save_lender_rates(rates: list) -> None:
    with _rates_lock:
        os.makedirs(os.path.dirname(RATES_FILE_PATH) or ".", exist_ok=True)
        with open(RATES_FILE_PATH, "w") as f:
            json.dump(rates, f, indent=2)


# ── EMI RULES ENGINE ────────────────────────────────────────────────────
def calculate_emi(principal: float, annual_rate: float, years: float) -> float:
    r = (annual_rate / 12) / 100
    n = years * 12
    if r == 0:
        return principal / n
    return (principal * r * (1 + r) ** n) / ((1 + r) ** n - 1)


def amortize(principal: float, annual_rate: float, years: float) -> dict:
    """
    Month-by-month amortization so we can answer the two questions borrowers
    never get shown at sanction: 'when does more than half my EMI start
    reducing MY loan instead of paying the bank's interest' (crossover point),
    and 'in year 1, how much of what I pay actually counts'.
    """
    r = (annual_rate / 12) / 100
    n = max(1, int(round(years * 12)))
    emi = calculate_emi(principal, annual_rate, years)

    balance = principal
    crossover_month = None
    yearly = []
    year_principal = 0.0
    year_interest = 0.0

    for m in range(1, n + 1):
        interest_component = balance * r
        principal_component = emi - interest_component
        balance = max(0.0, balance - principal_component)

        year_principal += principal_component
        year_interest += interest_component

        if crossover_month is None and principal_component > interest_component:
            crossover_month = m

        if m % 12 == 0 or m == n:
            yearly.append({
                "year": (m - 1) // 12 + 1,
                "principal_paid": round(year_principal),
                "interest_paid": round(year_interest),
            })
            year_principal = 0.0
            year_interest = 0.0

    first_year = yearly[0] if yearly else {"principal_paid": 0, "interest_paid": 0}
    first_year_total = first_year["principal_paid"] + first_year["interest_paid"]
    first_year_principal_pct = round(first_year["principal_paid"] / first_year_total * 100) if first_year_total else 0

    return {
        "crossover_month": crossover_month,
        "crossover_year": round(crossover_month / 12, 1) if crossover_month else None,
        "yearly_breakdown": yearly[:15],
        "first_year_principal_pct": first_year_principal_pct,
        "first_year_interest_pct": 100 - first_year_principal_pct,
    }


def build_calculation(principal: float, current_rate: float, years: float) -> dict:
    lender_rates = load_lender_rates()

    current_emi = calculate_emi(principal, current_rate, years)
    current_total_payment = current_emi * years * 12
    current_total_interest = current_total_payment - principal
    amort = amortize(principal, current_rate, years)

    best_lender = min(lender_rates, key=lambda l: l["rate"])
    new_emi = calculate_emi(principal, best_lender["rate"], years)
    new_total_payment = new_emi * years * 12
    total_savings = current_total_payment - new_total_payment
    monthly_savings = current_emi - new_emi

    rate_table = sorted(lender_rates, key=lambda l: l["rate"])
    rate_table_out = [
        {
            "name": l["name"],
            "rate": l["rate"],
            "emi": round(calculate_emi(principal, l["rate"], years)),
            "is_lowest": i == 0,
        }
        for i, l in enumerate(rate_table)
    ]

    # Plain-language line the frontend can show directly under the EMI number —
    # this is the single most important sentence banks never say out loud.
    if amort["crossover_year"]:
        crossover_plain = (
            f"For the first {amort['crossover_year']} years, more than half of every EMI you pay "
            f"is interest — it reduces the bank's risk, not your loan. Only after that point does "
            f"more than half of each EMI start reducing what you actually owe."
        )
    else:
        crossover_plain = (
            "Across this entire tenure, interest never drops below half of your EMI. "
            "A shorter tenure or a lower rate would change this."
        )

    return {
        "current_emi": round(current_emi),
        "current_total_interest": round(current_total_interest),
        "current_total_payment": round(current_total_payment),
        "current_rate": current_rate,
        "best_rate": best_lender["rate"],
        "best_lender_name": best_lender["name"],
        "total_savings": round(total_savings),
        "monthly_savings": round(monthly_savings),
        "worth_switching": total_savings > 0,
        "rate_table": rate_table_out,
        "amortization": amort,
        "crossover_plain_language": crossover_plain,
    }


# ── BORROWER EDUCATION CONTENT ──────────────────────────────────────────
# Static, deterministic, plain-language content — not AI-generated, so it
# never drifts and never needs an API key. Organised by the moment in the
# borrower's journey where each piece becomes relevant. Every section is
# written for someone taking their first loan, with rupee-level examples
# instead of jargon. Content is directional education, not legal advice —
# see the disclaimer field.

LOAN_EDUCATION_CONTENT = {
    "before_you_sign": {
        "title": "Before You Sign: What the Bank Won't Bring Up First",
        "points": [
            {
                "heading": "The bank earns from your interest, not your success",
                "body": (
                    "A bank's revenue is the interest you pay. The relationship manager's job is to "
                    "get you to sign, not to get you the cheapest loan. That's not dishonesty — it's "
                    "the business model. Your job is to shop the rate the same way you'd shop a phone price, "
                    "not accept the first number offered."
                ),
            },
            {
                "heading": "Ask for the total cost, not just the EMI",
                "body": (
                    "A lower EMI over a longer tenure can cost you far more in total interest. Before "
                    "signing, ask the bank in writing: 'What is the TOTAL amount I will pay over the "
                    "full tenure, including all interest?' Compare that number across lenders, not just the EMI."
                ),
            },
            {
                "heading": "'Insurance' bundled into the loan is usually optional",
                "body": (
                    "Many banks offer (and some push hard) a single-premium credit life insurance policy, "
                    "financed INTO your loan amount. This raises your principal, which raises your interest "
                    "for the entire tenure — insurance you're paying interest on for 20 years. Under IRDAI "
                    "and RBI guidance, loan-linked insurance is not mandatory for loan approval. You can say "
                    "'I do not want this added to my loan' and ask for a standalone term policy instead, "
                    "priced and bought separately, if you want life cover at all."
                ),
            },
            {
                "heading": "Processing fees, legal fees, and 'documentation charges' add up",
                "body": (
                    "Ask for a single sheet listing every charge: processing fee, technical/legal valuation "
                    "fee, stamp duty on the loan agreement, franking charges, CERSAI registration fee. These "
                    "are usually 0.5%–1.5% of the loan amount and are rarely volunteered upfront."
                ),
            },
            {
                "heading": "Check if you can prepay without penalty",
                "body": (
                    "For floating-rate home loans, RBI rules say banks CANNOT charge a prepayment or "
                    "foreclosure penalty to an individual borrower. If a bank official tells you otherwise "
                    "for a floating-rate loan, that is incorrect — ask them to show you where it's written "
                    "in the sanction letter."
                ),
            },
        ],
    },
    "while_repaying": {
        "title": "While You're Repaying: Why the First Years Feel Slow",
        "points": [
            {
                "heading": "Interest is calculated on what you still owe, every single month",
                "body": (
                    "In year one, your outstanding balance is close to the full loan amount, so the interest "
                    "portion of your EMI is large. As you slowly reduce the balance, the interest portion "
                    "shrinks and the principal portion grows — but this shift is gradual, not sudden. This is "
                    "why it can feel like 'nothing is moving' in the early years, even though every EMI is on time."
                ),
            },
            {
                "heading": "The crossover point is the milestone worth tracking",
                "body": (
                    "The month your EMI's principal component finally overtakes its interest component is "
                    "the real turning point. Use the calculator above — it's shown as 'crossover_year' — to "
                    "know when that happens for YOUR loan. Making even small prepayments early moves this "
                    "point closer, because it directly cuts the balance interest is calculated on."
                ),
            },
            {
                "heading": "A small extra payment early is worth more than a large one later",
                "body": (
                    "Because interest compounds on the outstanding balance, ₹50,000 paid extra in year 2 "
                    "saves more total interest than ₹50,000 paid extra in year 15. If you get a bonus or "
                    "windfall, part-prepaying a home loan in its early years is usually one of the highest, "
                    "safest returns available to a salaried borrower."
                ),
            },
        ],
    },
    "if_you_miss_a_payment": {
        "title": "If You Miss a Payment: The Real Timeline",
        "points": [
            {
                "heading": "Day 1–30: Grace period, but not a free pass",
                "body": (
                    "Most banks allow a short grace window, but you are typically charged a late payment fee "
                    "and penal interest on the overdue amount from day one. This already shows up on your "
                    "credit report the next reporting cycle, even a single missed EMI reported as overdue."
                ),
            },
            {
                "heading": "90 days overdue: Your loan becomes an NPA",
                "body": (
                    "If an EMI stays unpaid for 90 days, RBI rules require the bank to classify your loan as "
                    "a Non-Performing Asset (NPA). This is a formal, reported status — it is not just an "
                    "internal note. Once a loan is an NPA, the bank's recovery process becomes far more "
                    "aggressive and formal."
                ),
            },
            {
                "heading": "After that: Legal notice, then SARFAESI",
                "body": (
                    "For secured loans like a home loan, banks can invoke the SARFAESI Act after a 60-day "
                    "notice period post-NPA classification. This allows the bank to take possession of the "
                    "mortgaged property and sell it to recover dues — without needing to go to court first, "
                    "for standard cases above the notified threshold. This is the single biggest reason a "
                    "home loan default is treated far more seriously than a personal loan or credit card default."
                ),
            },
            {
                "heading": "Your CIBIL score takes the hit immediately, and it lingers",
                "body": (
                    "A single missed EMI, once reported, can drop a good CIBIL score noticeably. A settled or "
                    "written-off loan status stays visible on your credit report for up to 7 years, and can "
                    "block or worsen the terms of every future loan or credit card application in that window."
                ),
            },
            {
                "heading": "If you know a payment will be missed, call the bank before the due date",
                "body": (
                    "Banks have restructuring and moratorium options for genuine hardship, but these are "
                    "almost always easier to arrange BEFORE you default than after. Silence is the worst "
                    "option — it guarantees the formal recovery timeline above starts running."
                ),
            },
        ],
    },
    "if_you_are_a_guarantor": {
        "title": "If Someone Asks You to Be a Guarantor",
        "points": [
            {
                "heading": "A guarantor is not a character reference — it's equal liability",
                "body": (
                    "Signing as a guarantor means you are legally responsible for the full outstanding loan "
                    "if the borrower stops paying. The bank does not have to exhaust all options against the "
                    "borrower first — it can approach the guarantor directly for recovery."
                ),
            },
            {
                "heading": "It affects YOUR CIBIL score too",
                "body": (
                    "The guaranteed loan can appear on the guarantor's own credit report. If the borrower "
                    "misses payments, the guarantor's CIBIL score can drop, even though the guarantor never "
                    "missed a payment on their own loans."
                ),
            },
            {
                "heading": "It reduces YOUR own future borrowing capacity",
                "body": (
                    "Banks factor in guaranteed loan amounts when assessing a guarantor's own loan "
                    "eligibility later, since it counts as a contingent liability. Becoming a guarantor for a "
                    "large loan can quietly reduce how much you yourself can borrow in the future."
                ),
            },
            {
                "heading": "Before signing, ask these three questions",
                "body": (
                    "1) Can I see the borrower's repayment capacity and existing loans? "
                    "2) Am I comfortable paying the FULL EMI myself if the borrower cannot? "
                    "3) Is there a way to limit my guarantee to a specific amount or time period, in writing? "
                    "If the answer to question 2 is 'no', that is the clearest signal not to sign."
                ),
            },
        ],
    },
    "disclaimer": (
        "This content explains common rules and patterns (RBI/IRDAI norms, SARFAESI process, CIBIL "
        "reporting) in plain language for general understanding. It is not legal or financial advice, "
        "and specific terms vary by lender and loan agreement — always confirm against your own sanction "
        "letter and, for a live default or legal notice, consult a lawyer."
    ),
}


@app.route("/api/education", methods=["GET"])
def api_education():
    """
    Returns the full plain-language borrower-protection content. A `section`
    query param can request just one part, e.g.
    /api/education?section=if_you_miss_a_payment
    """
    section = request.args.get("section")
    if section:
        if section not in LOAN_EDUCATION_CONTENT:
            return jsonify({"error": f"Unknown section '{section}'.",
                             "available_sections": [k for k in LOAN_EDUCATION_CONTENT if k != "disclaimer"]}), 404
        return jsonify({section: LOAN_EDUCATION_CONTENT[section],
                         "disclaimer": LOAN_EDUCATION_CONTENT["disclaimer"]})
    return jsonify(LOAN_EDUCATION_CONTENT)


# ── ROUTES ───────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html", razorpay_key_id=RAZORPAY_KEY_ID)


@app.route("/health")
def health():
    return jsonify({"status": "ok"})


@app.route("/api/calculate-emi", methods=["POST"])
def api_calculate_emi():
    data = request.get_json(silent=True) or {}
    try:
        principal = float(data.get("loan_amount", 0))
        current_rate = float(data.get("current_rate", 0))
        years = float(data.get("tenure_years", 0))
    except (TypeError, ValueError):
        return jsonify({"error": "Invalid input. Please send numeric values."}), 400

    if principal <= 0 or current_rate <= 0 or years <= 0:
        return jsonify({"error": "loan_amount, current_rate, and tenure_years must all be positive numbers."}), 400

    result = build_calculation(principal, current_rate, years)
    return jsonify(result)


@app.route("/api/create-order", methods=["POST"])
def api_create_order():
    if not razorpay_client:
        return jsonify({"error": "Payments are not configured on this server yet."}), 503

    data = request.get_json(silent=True) or {}
    notes = {
        "product": "home-loan-manager-report",
        "loan_amount": str(data.get("loan_amount", "")),
        "current_rate": str(data.get("current_rate", "")),
        "tenure_years": str(data.get("tenure_years", "")),
    }

    try:
        order = razorpay_client.order.create(
            {
                "amount": REPORT_PRICE_PAISE,
                "currency": "INR",
                "receipt": f"hlm_{os.urandom(6).hex()}",
                "notes": notes,
                "payment_capture": 1,
            }
        )
    except Exception as exc:
        logger.exception("Razorpay order creation failed")
        return jsonify({"error": "Could not create payment order.", "detail": str(exc)}), 502

    return jsonify(
        {
            "order_id": order["id"],
            "amount": order["amount"],
            "currency": order["currency"],
            "key_id": RAZORPAY_KEY_ID,
        }
    )


@app.route("/api/verify-payment", methods=["POST"])
def api_verify_payment():
    if not razorpay_client:
        return jsonify({"error": "Payments are not configured on this server yet."}), 503

    data = request.get_json(silent=True) or {}
    params = {
        "razorpay_order_id": data.get("razorpay_order_id", ""),
        "razorpay_payment_id": data.get("razorpay_payment_id", ""),
        "razorpay_signature": data.get("razorpay_signature", ""),
    }

    if not all(params.values()):
        return jsonify({"error": "Missing Razorpay verification fields."}), 400

    try:
        razorpay_client.utility.verify_payment_signature(params)
    except razorpay.errors.SignatureVerificationError:
        return jsonify({"error": "Payment signature verification failed.", "verified": False}), 400

    lead = data.get("lead", {})
    calc = data.get("calc", {})
    lead_record = {
        "name": lead.get("name", ""),
        "phone": lead.get("phone", ""),
        "email": lead.get("email", ""),
        "city": lead.get("city", ""),
        "loan_amount": calc.get("loan_amount", ""),
        "current_rate": calc.get("current_rate", ""),
        "tenure_years": calc.get("tenure_years", ""),
        "best_lender": calc.get("best_lender_name", ""),
        "total_savings": calc.get("total_savings", ""),
        "payment_id": params["razorpay_payment_id"],
        "order_id": params["razorpay_order_id"],
        "amount_paid_inr": REPORT_PRICE_PAISE / 100,
        "product": "home-loan-manager-report",
        "status": "paid",
    }

    sheet_ok = push_lead_to_sheet(lead_record)

    return jsonify({"verified": True, "lead_saved": sheet_ok})


@app.route("/api/lead-capture", methods=["POST"])
def api_lead_capture():
    data = request.get_json(silent=True) or {}
    name = (data.get("name") or "").strip()
    phone = (data.get("phone") or "").strip()

    if not name or not phone:
        return jsonify({"error": "name and phone are required."}), 400

    lead_record = {
        "name": name,
        "phone": phone,
        "email": (data.get("email") or "").strip(),
        "city": (data.get("city") or "").strip(),
        "loan_amount": data.get("loan_amount", ""),
        "current_rate": data.get("current_rate", ""),
        "tenure_years": data.get("tenure_years", ""),
        "payment_id": "",
        "order_id": "",
        "amount_paid_inr": 0,
        "product": "home-loan-manager-lead",
        "status": "unpaid_lead",
    }

    sheet_ok = push_lead_to_sheet(lead_record)
    return jsonify({"saved": sheet_ok})


def _get_or_create_workbook():
    if os.path.exists(LEADS_FILE_PATH):
        return load_workbook(LEADS_FILE_PATH)
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    ws.append(LEADS_HEADERS)
    return wb


def push_lead_to_sheet(record: dict) -> bool:
    row = [datetime.datetime.utcnow().isoformat()] + [record.get(h, "") for h in LEADS_HEADERS[1:]]

    try:
        with _leads_lock:
            os.makedirs(os.path.dirname(LEADS_FILE_PATH) or ".", exist_ok=True)
            wb = _get_or_create_workbook()
            ws = wb["Leads"] if "Leads" in wb.sheetnames else wb.active
            ws.append(row)
            wb.save(LEADS_FILE_PATH)
        return True
    except Exception:
        logger.exception("Failed to append lead to Excel file at %s", LEADS_FILE_PATH)
        return False


@app.route("/admin/rates", methods=["GET", "POST"])
def admin_rates():
    if not ADMIN_TOKEN:
        abort(503, description="ADMIN_TOKEN is not configured on this server.")
    token = request.args.get("token") or request.form.get("token")
    if token != ADMIN_TOKEN:
        abort(403)

    message = ""
    if request.method == "POST":
        names = request.form.getlist("name")
        rates = request.form.getlist("rate")
        new_rates = []
        try:
            for name, rate in zip(names, rates):
                name = name.strip()
                if not name:
                    continue
                new_rates.append({"name": name, "rate": float(rate)})
            if not new_rates:
                raise ValueError("At least one lender row is required.")
            save_lender_rates(new_rates)
            message = "Rates updated successfully."
        except ValueError as exc:
            message = f"Error: {exc}"

    current_rates = load_lender_rates()

    rows_html = "".join(
        f"""
        <tr>
          <td><input type="text" name="name" value="{l['name']}" style="width:220px;padding:6px"></td>
          <td><input type="number" step="0.01" name="rate" value="{l['rate']}" style="width:90px;padding:6px"></td>
        </tr>
        """
        for l in current_rates
    )

    blank_rows = "".join(
        """
        <tr>
          <td><input type="text" name="name" value="" style="width:220px;padding:6px" placeholder="Lender name"></td>
          <td><input type="number" step="0.01" name="rate" value="" style="width:90px;padding:6px" placeholder="Rate %"></td>
        </tr>
        """
        for _ in range(3)
    )

    html = f"""
    <html>
    <head><title>Home Loan Manager — Update Rates</title>
    <style>
      body {{ font-family: sans-serif; max-width: 600px; margin: 40px auto; color: #1a1a2e; }}
      h1 {{ font-size: 20px; }}
      table {{ border-collapse: collapse; margin: 20px 0; }}
      th, td {{ padding: 6px 10px; text-align: left; }}
      button {{ background: #5b6af0; color: #fff; border: none; padding: 10px 20px; border-radius: 6px; font-size: 14px; cursor: pointer; }}
      .msg {{ padding: 10px; border-radius: 6px; background: #eeeef6; margin-bottom: 10px; }}
    </style>
    </head>
    <body>
      <h1>Lender Rates</h1>
      <p>Update starting rates here whenever RBI revises the repo rate or a DSA partner flags a change. No redeploy needed — saved instantly.</p>
      {'<div class="msg">' + message + '</div>' if message else ''}
      <form method="POST">
        <input type="hidden" name="token" value="{token}">
        <table>
          <tr><th>Lender Name</th><th>Rate (%)</th></tr>
          {rows_html}
          {blank_rows}
        </table>
        <button type="submit">Save Rates</button>
      </form>
    </body>
    </html>
    """
    return Response(html, mimetype="text/html")


@app.route("/admin/leads")
def admin_download_leads():
    if not ADMIN_TOKEN:
        abort(503, description="ADMIN_TOKEN is not configured on this server.")
    if request.args.get("token") != ADMIN_TOKEN:
        abort(403)
    if not os.path.exists(LEADS_FILE_PATH):
        abort(404, description="No leads recorded yet.")
    return send_file(LEADS_FILE_PATH, as_attachment=True, download_name="leads.xlsx")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
